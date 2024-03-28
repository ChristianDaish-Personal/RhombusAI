# This python script processes chunks for large datasets
# It was developed by Christian Daish on the 26th of March 2024 for Rhombus AI.
# This is version 1.0.260324
#Import packages and functions
from .infer_and_convert_data import infer_and_convert_dtypes
from .dataload import load_data
from openpyxl import load_workbook
import pandas as pd

# Process the file correctly and perform chunking where filesize is large, then send to be inferred and converted.
def process_files(file_path,file_type,chunksize=None, threshold=0.5): #default values for chunksize and threshold
    data_or_iterator, data_type = load_data(file_path,file_type,chunksize)

    if data_type == 'csv_chunked':
        # Process each chunk for CSV files
        for chunk in data_or_iterator:
            converted_chunk = infer_and_convert_dtypes(chunk, threshold)
            pass
            # Process each converted_chunk here
            result = converted_chunk.dtypes
    elif data_type == 'xlsx_chunked':
        for segment in load_excel_in_segments(file_path, chunksize):
            converted_segment = infer_and_convert_dtypes(segment, threshold)
            pass
            # Process complete file here
            result = converted_segment.dtypes
    elif data_type in ('csv_full','xlsx_full'):
        converted_df = infer_and_convert_dtypes(data_or_iterator, threshold)
        pass
        # Process complete file here
        result = converted_df.dtypes
    return file_path,result

#function for counting rows in .xlsx files
def get_excel_file_row_count(file_path):
    """Get the total number of rows in the Excel file, excluding the header."""
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active
    return sheet.max_row - 1

#function for segmentation/chunking of .xlsx files
def load_excel_in_segments(file_path, segment_size):
    """Generator function to load Excel file in segments."""
    total_rows = get_excel_file_row_count(file_path)
    start_row = 0  # Start at the first row after the header

    while start_row < total_rows:
        df_segment = pd.read_excel(file_path, skiprows=range(1, start_row + 1), nrows=segment_size)
        yield df_segment
        start_row += segment_size
